"""
导出服务 - Excel导出功能
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from models.attendance import AttendanceRecord
from models.student import Student
from models.course import Course

class ExportService:
    """导出服务"""

    @staticmethod
    def export_attendance_to_excel(course_id, records, output_path=None):
        """
        导出考勤记录到Excel
        """
        if output_path is None:
            output_path = os.path.join(os.path.dirname(__file__), '..', 'exports')
            os.makedirs(output_path, exist_ok=True)
            filename = f"attendance_{course_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(output_path, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "考勤记录"

        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ["学号", "姓名", "课程", "签到时间", "签到方式", "考勤状态", "置信度", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.student.student_code if record.student else "")
            ws.cell(row=row, column=2, value=record.student.name if record.student else "")
            ws.cell(row=row, column=3, value=record.course.course_name if record.course else "")
            ws.cell(row=row, column=4, value=record.check_in_time.strftime("%Y-%m-%d %H:%M:%S") if record.check_in_time else "")
            ws.cell(row=row, column=5, value="摄像头签到" if record.check_in_type == "camera" else "手动签到")
            ws.cell(row=row, column=6, value=ExportService._get_status_text(record.status))
            ws.cell(row=row, column=7, value=f"{float(record.confidence):.2%}" if record.confidence else "")
            ws.cell(row=row, column=8, value=record.remarks or "")

            # 设置边框
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border

        # 设置列宽
        column_widths = [15, 10, 20, 20, 12, 10, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(output_path)
        return output_path

    @staticmethod
    def _get_status_text(status):
        """获取状态文本"""
        status_map = {
            'normal': '正常',
            'late': '迟到',
            'absent': '缺勤',
            'leave': '请假'
        }
        return status_map.get(status, status)

    @staticmethod
    def export_statistics_to_excel(course_id, statistics, students, output_path=None):
        """
        导出考勤统计到Excel
        """
        if output_path is None:
            output_path = os.path.join(os.path.dirname(__file__), '..', 'exports')
            os.makedirs(output_path, exist_ok=True)
            filename = f"statistics_{course_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(output_path, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "考勤统计"

        # 样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = ["学号", "姓名", "正常次数", "迟到次数", "缺勤次数", "请假次数", "出勤率"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 统计数据
        stats_dict = {s['student_id']: s for s in statistics}

        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student.student_code)
            ws.cell(row=row, column=2, value=student.name)

            if student.id in stats_dict:
                s = stats_dict[student.id]
                ws.cell(row=row, column=3, value=s.get('normal', 0))
                ws.cell(row=row, column=4, value=s.get('late', 0))
                ws.cell(row=row, column=5, value=s.get('absent', 0))
                ws.cell(row=row, column=6, value=s.get('leave', 0))
                ws.cell(row=row, column=7, value=f"{s.get('attendance_rate', 0):.1f}%")
            else:
                ws.cell(row=row, column=3, value=0)
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=0)
                ws.cell(row=row, column=7, value="0.0%")

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border

        # 列宽
        column_widths = [15, 10, 12, 12, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(output_path)
        return output_path
